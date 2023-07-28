from flask import Flask, render_template, request, redirect, url_for, send_file
import os
import openpyxl
from openpyxl import Workbook
from datetime import datetime


app = Flask(__name__)


BASE_DIR = os.path.abspath(os.path.dirname(__file__))
EXCEL_FILE_PATH = os.path.join(BASE_DIR, 'data.xlsx')

if not os.path.exists(EXCEL_FILE_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Total", "1 centime", "2 centimes", "5 centimes", "10 centimes", "20 centimes",
               "50 centimes", "1 euro", "2 euros", "5 euros", "10 euros", "20 euros", "50 euros",
               "100 euros", "200 euros", "Enlevé 5 euros", "Enlevé 10 euros", "Enlevé 20 euros", "Enlevé 50 euros", "Enlevé 100 euros", "Enlevé 200 euros"])
    wb.save(EXCEL_FILE_PATH)

def add_to_excel(data):
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = rows[0]
    ldata = rows[1:]

    did = False
    for i in range(len(ldata)):
        row = ldata[i]
        print("row", row)
        print("data:", data)
        if row[0] == data[0]:
            ldata[i] = data
            did = True
            break  # Ajoutez cette ligne pour sortir de la boucle une fois que la ligne a été trouvée et modifiée

    if not did:
        ws.append(data)

    # Réécrire toutes les données (y compris la ligne modifiée, si elle existe) dans le fichier Excel
    all_data = [headers] + ldata
    for row_idx, row_data in enumerate(all_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

    wb.save(EXCEL_FILE_PATH)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date_str = request.form['date']

        n = ["1", "2", "5", "10", "20", "50", "1_euro", "2_euros"]
        nb_coins = []
        coins = []
        for i in n:
            coin_key = f'coin_{i}'
            if "_" in i:
                if "s" in i:
                    try :
                        nb = int(request.form.get(coin_key, 0))
                        coin_value = 2 * nb
                    except:
                        nb = 0
                        coin_value = 0
                else:
                    try:
                        nb = int(request.form.get(coin_key, 0))
                        coin_value = nb
                    except:
                        nb = 0
                        coin_value = 0
            else:
                try:
                    nb = int(request.form.get(coin_key, 0))
                    coin_value = int(i) * nb * 0.01
                except:
                    nb = 0
                    coin_value = 0
            coins.append(coin_value)
            nb_coins.append(nb)

        n = ["5", "10", "20", "50", "100", "200"]
        nb_banknotes = []
        banknotes = []
        for i in n:
            banknote_key = f'banknote_{i}'
            try:
                nb = int(request.form.get(banknote_key, 0))
                banknote_value = nb * int(i)
            except:
                nb = 0
                banknote_value = 0
            banknotes.append(banknote_value)
            nb_banknotes.append(nb)
        
        delbanknotes = []
        nb_delbanknotes = []
        for i in n:
            delbanknote_key = f'delbanknote_{i}'
            try:
                nb = int(request.form.get(delbanknote_key, 0))
                delbanknote_value = nb * int(i)
            except:
                nb = 0
                delbanknote_value = 0
            delbanknotes.append(delbanknote_value)
            nb_delbanknotes.append(nb)
        total = sum(coins) + sum(banknotes) - sum(delbanknotes)

        data = [date_str] + [total]  + nb_coins + nb_banknotes + nb_delbanknotes

        add_to_excel(data)

        return redirect(url_for('summary'))

    return render_template('index.html')


@app.route('/summary')
def summary():
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = rows[0]
    data = rows[1:]

    summary_data = []
    for row in data:
        summary_data.append(dict(zip(headers, row)))

    return render_template('summary.html', data=summary_data)

def excel_to_html_mapping(excel_value):
    mapping = {
        "Date" : "date",
        "1 centime": "coin_1",
        "2 centimes": "coin_2",
        "5 centimes": "coin_5",
        "10 centimes": "coin_10",
        "20 centimes": "coin_20",
        "50 centimes": "coin_50",
        "1 euro": "coin_1_euro",
        "2 euros": "coin_2_euros",
        "5 euros": "banknote_5",
        "10 euros": "banknote_10",
        "20 euros": "banknote_20",
        "50 euros": "banknote_50",
        "100 euros": "banknote_100",
        "200 euros": "banknote_200",
        "Enlevé 5 euros": "delbanknote_5",
        "Enlevé 10 euros": "delbanknote_10",
        "Enlevé 20 euros": "delbanknote_20",
        "Enlevé 50 euros": "delbanknote_50",
        "Enlevé 100 euros": "delbanknote_100",
        "Enlevé 200 euros": "delbanknote_200",
    }
    return mapping.get(excel_value, excel_value)


def html_to_excel_mapping(html_value):
    mapping = {v: k for k, v in excel_to_html_mapping({}).items()}
    return mapping.get(html_value, html_value)

@app.route('/precise_date',  methods=['POST'])
def precise_date():
    precise_date_str = request.form['date']

    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    headers = rows[0]
    data = rows[1:]

    precise_data = None
    for row in data:
        if row[0] == precise_date_str:
            precise_data = dict(zip(headers, row))
            break
    new_precise_data = {}
    for (k,v) in precise_data.items():
        new_precise_data[excel_to_html_mapping(k)] = v

    if precise_data:
        return render_template('index.html', precise_data=new_precise_data)
    else:
        # Handle case when no data is found for the specific date
        return render_template('index.html')
    

@app.route('/get_file', methods=['POST'])
def get_file():
    return send_file(EXCEL_FILE_PATH)

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5002)